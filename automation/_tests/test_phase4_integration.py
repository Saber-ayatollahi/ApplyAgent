"""Phase 4 integration test — end-to-end consistency across artifacts.

Per implementation plan § Phase 4: "full pipeline run with one sector mute
+ one manual selection; verify all artifacts (worklist_scored.json,
promote_report, audit pack xlsx, snapshot json) carry consistent state."

We don't run the actual scrape or LLM — those need network + API keys. We
synthesize a worklist_scored.json, drop a sector mute, then drive
auto_promote.main() and audit_pack.build_audit_pack() against it. The
test asserts:

  1. The promote_report shows the muted row in suppressed_after_score
     (because we didn't pass --only-urls — bulk threshold flow respects
     mutes).
  2. A second run with --only-urls including the same muted URL promotes
     it WITH selection_mode=manual_override_suppression.
  3. The audit pack xlsx surfaces both states across its sheets.

This is the load-bearing cross-cluster contract from the v3.1.1 spec.
"""
from __future__ import annotations

import io
import json
import os
import shutil
import subprocess
import sys
import tempfile
from datetime import date, timedelta
from pathlib import Path

import pytest

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent.parent
sys.path.insert(0, str(ROOT))


@pytest.fixture
def isolated_repo(monkeypatch, tmp_path):
    """Build an isolated copy of the directory layout the modules read.

    We don't isolate auto_promote globally — its OUT_DIR / TRACKER paths
    are module-level constants resolved at import. Instead we copy the
    real automation/ files we need, place them under tmp_path, and
    monkeypatch the path globals that matter."""
    # Layout:
    #   tmp_path/
    #     data/job_tracker_data.json
    #     data/suppressions.json + ancillary
    #     automation/outputs/worklist_scored.json
    #     automation/outputs/promote_report_*.json + .md
    data_dir = tmp_path / "data"
    out_dir = tmp_path / "automation" / "outputs"
    pipelines_dir = out_dir / "pipelines"
    data_dir.mkdir(parents=True)
    out_dir.mkdir(parents=True)
    pipelines_dir.mkdir(parents=True)

    # Seed an empty tracker.
    tracker_path = data_dir / "job_tracker_data.json"
    tracker_path.write_text(json.dumps({
        "meta": {"schema_version": 3,
                 "status_enum": ["Found", "Watch", "Tailoring", "Applied",
                                  "Recruiter_Screen", "Phone_Screen",
                                  "Take_Home", "Onsite", "Offer", "Rejected",
                                  "Withdrawn", "Hired", "Expired"]},
        "jobs": [],
    }, indent=2), encoding="utf-8")

    # Seed an empty suppressions registry. The CLI / module will lazy-create.
    (data_dir / "suppressions.example.json").write_text(
        json.dumps({"version": 1, "sectors": [], "companies": []}),
        encoding="utf-8",
    )

    # Patch every module's path globals to point at this temp tree.
    from automation import auto_promote
    from automation import suppressions as supp
    from automation import audit_pack
    monkeypatch.setattr(auto_promote, "OUT_DIR", out_dir)
    monkeypatch.setattr(auto_promote, "TRACKER", tracker_path)
    monkeypatch.setattr(supp, "LIVE_PATH", data_dir / "suppressions.json")
    monkeypatch.setattr(supp, "EXAMPLE_PATH", data_dir / "suppressions.example.json")
    monkeypatch.setattr(supp, "EVENTS_PATH", data_dir / "suppressions_events.jsonl")
    monkeypatch.setattr(supp, "HISTORY_PATH", data_dir / "suppressions_history.json")
    monkeypatch.setattr(supp, "PENDING_ARCHIVES_PATH",
                        data_dir / "suppressions_pending_archives.jsonl")
    monkeypatch.setattr(audit_pack, "OUT_DIR", out_dir)

    return {
        "tmp": tmp_path, "out_dir": out_dir, "data_dir": data_dir,
        "tracker": tracker_path,
    }


def _build_scored_envelope(scan_date: str) -> dict:
    """Two rows above the threshold — one in a soon-to-be-muted sector,
    one outside. Stage-1 carries no drops (we want the suppression to
    happen at promote time / via --only-urls, not at triage)."""
    return {
        "scan_date": scan_date,
        "stage1_passed": 2,
        "stage1_dropped": 0,
        "stage2_scored": 2,
        "results": [
            {"title": "Director, ALM", "company": "RBC",
             "sector": "Canadian Big 6 Banks",
             "location": "Toronto, ON",
             "source": "scrape", "link": "https://x.co/rbc-alm",
             "fit": {"fit_score": 9, "fit_verdict": "apply_now",
                     "tier": 1, "summary": "great",
                     "top_3_reasons": ["alm match"], "skill_gaps": []}},
            {"title": "VP Treasury", "company": "Wave",
             "sector": "Fintech",
             "location": "Toronto, ON",
             "source": "scrape", "link": "https://x.co/wave-vp",
             "fit": {"fit_score": 8, "fit_verdict": "apply_now",
                     "tier": 2, "summary": "ok",
                     "top_3_reasons": ["treasury match"], "skill_gaps": []}},
        ],
        "triage_drops": [],
    }


def test_phase4_full_loop_threshold_then_manual_override(isolated_repo):
    """The integration scenario: bulk threshold respects the mute; manual
    selection of the muted URL overrides it; the audit pack surfaces both."""
    from automation import auto_promote
    from automation import suppressions as supp
    from automation import audit_pack

    scan_date = "29991231"
    scored = _build_scored_envelope(scan_date)
    scored_path = isolated_repo["out_dir"] / "worklist_scored.json"
    scored_path.write_text(json.dumps(scored), encoding="utf-8")

    # Add a sector mute. Going through the real API so canonical-key
    # matching is exercised.
    until = date.today() + timedelta(days=60)
    supp.add_sector("Canadian Big 6 Banks", until,
                     "Q2 cooldown — integration test")
    active = supp.load_active()
    assert any(e["canonical_key"] == "canadian big 6 banks"
               for e in active.get("sectors", []))

    # ── Pass 1: threshold flow (no --only-urls) ──
    # Bulk threshold should drop the muted RBC row into suppressed_after_score.
    rc1 = auto_promote.main_with_args([
        "--scan", scored_path.name,
        "--commit", "--min-score", "7",
    ]) if hasattr(auto_promote, "main_with_args") else None
    if rc1 is None:
        # Fall back to argv injection
        old_argv = sys.argv
        sys.argv = ["auto_promote.py", "--scan", scored_path.name,
                    "--commit", "--min-score", "7"]
        try:
            rc1 = auto_promote.main()
        finally:
            sys.argv = old_argv
    assert rc1 == 0, f"threshold pass exit code: {rc1}"

    # Find the promote report. auto_promote writes
    # promote_report_<scan_date>.json next to the scored file.
    reports = list(isolated_repo["out_dir"].glob("promote_report_*.json"))
    assert len(reports) >= 1, "no promote report after threshold pass"
    threshold_report = json.loads(reports[0].read_text(encoding="utf-8"))

    # The muted RBC row must NOT be in new_entries; it must be in
    # suppressed_after_score with promoted_anyway=False.
    new_urls = {e.get("url") for e in threshold_report.get("new_entries", []) or []}
    sup_rows = threshold_report.get("suppressed_after_score") or []
    assert "https://x.co/rbc-alm" not in new_urls, (
        "threshold flow promoted a muted row")
    rbc_sup = next((r for r in sup_rows
                    if r.get("url") == "https://x.co/rbc-alm"), None)
    assert rbc_sup is not None, (
        "muted row missing from suppressed_after_score")
    assert not rbc_sup.get("promoted_anyway", False)

    # The unmuted Wave row should have been promoted.
    assert "https://x.co/wave-vp" in new_urls

    # ── Pass 2: manual selection override of the muted URL ──
    # User explicitly picks the RBC URL despite the active mute.
    picks_file = isolated_repo["out_dir"] / "manual_picks.txt"
    picks_file.write_text("https://x.co/rbc-alm\n", encoding="utf-8")

    old_argv = sys.argv
    sys.argv = ["auto_promote.py",
                "--scan", scored_path.name,
                "--commit",
                "--only-urls", str(picks_file)]
    try:
        rc2 = auto_promote.main()
    finally:
        sys.argv = old_argv
    assert rc2 == 0, f"manual override exit code: {rc2}"

    # The latest report should now include the RBC row with
    # selection_mode=manual_override_suppression.
    reports = sorted(isolated_repo["out_dir"].glob("promote_report_*.json"),
                     key=lambda p: p.stat().st_mtime, reverse=True)
    manual_report = json.loads(reports[0].read_text(encoding="utf-8"))
    new_entries = manual_report.get("new_entries", []) or []
    rbc_promoted = next((e for e in new_entries
                         if e.get("url") == "https://x.co/rbc-alm"), None)
    assert rbc_promoted is not None, (
        "manual override didn't promote the muted URL")
    assert rbc_promoted.get("selection_mode") == "manual_override_suppression"
    # Run-level selection_mode reflects the manual flow.
    assert manual_report.get("selection_mode") in ("manual", "mixed")

    # ── Audit pack consistency. Per auto_promote convention the report
    # stamp uses date.today() not the scan_date in the envelope, so
    # resolve via the actual filename pattern.
    promote_stamp = reports[0].stem.replace("promote_report_", "")
    pack_bytes = audit_pack.build_audit_pack(promote_stamp)
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(pack_bytes), read_only=True)
    assert "Suppressed (race)" in wb.sheetnames
    assert "Promoted" in wb.sheetnames
    # Promoted sheet must include the override row with the selection_mode
    promoted_ws = wb["Promoted"]
    rows = list(promoted_ws.iter_rows(values_only=True))
    headers = list(rows[0])
    sm_idx = headers.index("selection_mode") if "selection_mode" in headers else None
    assert sm_idx is not None, "selection_mode column missing from Promoted sheet"
    promoted_data = [dict(zip(headers, r)) for r in rows[1:]]
    rbc_in_pack = next((r for r in promoted_data
                        if r.get("url") == "https://x.co/rbc-alm"), None)
    assert rbc_in_pack is not None
    assert rbc_in_pack.get("selection_mode") == "manual_override_suppression"
