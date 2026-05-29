"""Test audit_pack.build_audit_pack() against synthetic envelopes.

Verifies:
  - Multi-sheet xlsx round-trips through openpyxl with all expected sheets.
  - Each sheet's row count matches the input envelope.
  - Empty drop-logs (legacy scans pre-instrumentation) still produce sheets.
"""
from __future__ import annotations

import io
import json
import os
import sys
import tempfile
from datetime import datetime
from pathlib import Path

HERE = Path(__file__).resolve().parent
AUTO = HERE.parent
sys.path.insert(0, str(AUTO))


def _make_synth_outputs(out_dir: Path, stamp: str) -> None:
    """Write minimal scan/worklist/scored/promote envelopes the audit_pack
    builder reads from. Each envelope has 2-3 rows so we can verify sheet
    row counts deterministically."""
    out_dir.mkdir(parents=True, exist_ok=True)

    scan = {
        "scan_date": stamp,
        "companies_scanned": 2,
        "total_new_candidates": 2,
        "dedup_stats": {"input": 3, "output": 2, "dropped_url": 1, "dropped_near": 0},
        "by_sector": {"Banks": 2},
        "diagnostics": {},
        "filter_drops": {
            "title": [
                {"company": "Acme", "sector": "Banks",
                 "title": "Marketing Intern", "link": "https://x.co/1",
                 "location": "Toronto", "source": "workday:acme",
                 "matched_terms": "intern,marketing"},
            ],
            "geo": [
                {"company": "Acme", "sector": "Banks",
                 "title": "Risk Analyst", "link": "https://x.co/2",
                 "location": "Boston, MA", "source": "greenhouse:acme"},
                {"company": "Acme", "sector": "Banks",
                 "title": "Quant Analyst", "link": "https://x.co/3",
                 "location": "NYC", "source": "linkedin"},
            ],
        },
        "results": [
            {"title": "Risk Manager", "link": "https://x.co/10",
             "location": "Toronto, ON", "source": "workday:acme",
             "company": "Acme", "sector": "Banks"},
            {"title": "ALM Analyst", "link": "https://x.co/11",
             "location": "Toronto, ON", "source": "linkedin",
             "company": "Beta", "sector": "Banks"},
        ],
    }
    (out_dir / f"scan_{stamp}.json").write_text(
        json.dumps(scan), encoding="utf-8")

    worklist = {
        "version": 1,
        "scan_date": stamp,
        "stats": {"total": 2},
        "merged_pairs": [
            {"kept_url": "https://x.co/10", "kept_source": "scrape",
             "dropped_url": "https://x.co/10b", "dropped_source": "gmail",
             "company": "Acme", "title": "Risk Manager",
             "reason": "near_dup_company_title"},
        ],
        "results": [
            {"title": "Risk Manager", "link": "https://x.co/10",
             "company": "Acme", "sector": "Banks", "source": "scrape",
             "first_seen": stamp, "in_pool_since": stamp,
             "is_new_since_last_score": True},
            {"title": "ALM Analyst", "link": "https://x.co/11",
             "company": "Beta", "sector": "Banks", "source": "scrape",
             "first_seen": stamp, "in_pool_since": stamp,
             "is_new_since_last_score": True},
        ],
    }
    (out_dir / "worklist.json").write_text(
        json.dumps(worklist), encoding="utf-8")

    scored = {
        "scan_date": stamp,
        "stage1_passed": 2,
        "stage1_dropped": 1,
        "results": [
            {"title": "Risk Manager", "company": "Acme", "location": "Toronto",
             "source": "scrape", "link": "https://x.co/10",
             "fit": {"fit_verdict": "tailor_and_apply", "fit_score": 78,
                     "fit_notes": "good ALM match"}},
            {"title": "ALM Analyst", "company": "Beta", "location": "Toronto",
             "source": "scrape", "link": "https://x.co/11",
             "fit": {"fit_verdict": "watch", "fit_score": 55, "fit_notes": ""}},
        ],
        "triage_drops": [
            {"company": "Gamma", "title": "Sales Rep",
             "link": "https://x.co/20", "source": "linkedin",
             "rule_reasons": ["no_strong_keywords"], "score": 1,
             "hits_breakdown": {"strong": [], "medium": [], "weak": [],
                                "level": []}},
        ],
        "only_filtered": [],
    }
    (out_dir / "worklist_scored.json").write_text(
        json.dumps(scored), encoding="utf-8")

    promote = {
        "scan_date": stamp,
        "summary": {"added": 1, "skipped_geo": 1, "skipped_score": 0,
                    "skipped_verdict": 0, "skipped_dupe": 0},
        "skipped_rows": [
            {"reason": "geo", "company": "Acme", "title": "Risk Manager",
             "url": "https://x.co/99", "location": "NYC",
             "score": 70, "verdict": "tailor_and_apply", "note": "NYC"},
        ],
        "new_entries": [
            {"id": "auto-acme-risk-78", "tier": 1, "score": 78,
             "sector": "Banks", "company": "Acme", "title": "Risk Manager",
             "url": "https://x.co/10"},
        ],
    }
    (out_dir / f"promote_report_{stamp}.json").write_text(
        json.dumps(promote), encoding="utf-8")
    (out_dir / f"promote_report_{stamp}.md").write_text(
        "# stub", encoding="utf-8")


def test_audit_pack_round_trip():
    """Build a pack from synthetic envelopes, parse it back, check sheets."""
    import audit_pack  # type: ignore
    from openpyxl import load_workbook

    with tempfile.TemporaryDirectory() as tmp:
        synth_out = Path(tmp) / "outputs"
        _make_synth_outputs(synth_out, "29991231")

        # Point audit_pack at our temp dir.
        original_out = audit_pack.OUT_DIR
        audit_pack.OUT_DIR = synth_out
        try:
            data = audit_pack.build_audit_pack("29991231")
            assert isinstance(data, bytes) and len(data) > 1000, "tiny pack"

            wb = load_workbook(io.BytesIO(data), read_only=True)
            sheets = wb.sheetnames
            expected = {
                "Summary", "Scrape raw", "Scrape title drops",
                "Scrape geo drops", "Gmail raw", "Worklist pool",
                "Worklist merges", "Stage-1 drops", "Scored",
                "Promote skips", "Promoted",
            }
            missing = expected - set(sheets)
            assert not missing, f"missing sheets: {missing}"

            def rows(name: str) -> int:
                ws = wb[name]
                return ws.max_row - 1  # exclude header

            assert rows("Scrape raw") == 2, rows("Scrape raw")
            assert rows("Scrape title drops") == 1, rows("Scrape title drops")
            assert rows("Scrape geo drops") == 2, rows("Scrape geo drops")
            assert rows("Worklist pool") == 2, rows("Worklist pool")
            assert rows("Worklist merges") == 1, rows("Worklist merges")
            assert rows("Stage-1 drops") == 1, rows("Stage-1 drops")
            assert rows("Scored") == 2, rows("Scored")
            assert rows("Promote skips") == 1, rows("Promote skips")
            assert rows("Promoted") == 1, rows("Promoted")
            print(f"[PASS] audit_pack: all {len(expected)} sheets present, "
                  f"row counts match")
        finally:
            audit_pack.OUT_DIR = original_out


def test_audit_pack_empty_legacy_scan():
    """A scan envelope without `filter_drops` (legacy, pre-instrumentation)
    should still build a pack with empty title/geo drop sheets."""
    import audit_pack  # type: ignore
    from openpyxl import load_workbook

    with tempfile.TemporaryDirectory() as tmp:
        synth_out = Path(tmp) / "outputs"
        synth_out.mkdir(parents=True, exist_ok=True)
        legacy = {
            "scan_date": "29991230",
            "companies_scanned": 1,
            "results": [
                {"title": "T", "link": "https://x.co/1", "company": "A",
                 "sector": "B", "source": "scrape", "location": "Toronto"},
            ],
        }
        (synth_out / "scan_29991230.json").write_text(
            json.dumps(legacy), encoding="utf-8")
        original_out = audit_pack.OUT_DIR
        audit_pack.OUT_DIR = synth_out
        try:
            data = audit_pack.build_audit_pack("29991230")
            wb = load_workbook(io.BytesIO(data), read_only=True)
            assert "Scrape title drops" in wb.sheetnames
            assert "Scrape geo drops" in wb.sheetnames
            assert wb["Scrape raw"].max_row - 1 == 1
            assert wb["Scrape title drops"].max_row - 1 == 0
            print("[PASS] audit_pack: legacy scan envelope handled")
        finally:
            audit_pack.OUT_DIR = original_out


# ---------------------------------------------------------------------------
# Phase 4 — suppression columns + sheets, selection_mode column, run-meta sheet
# ---------------------------------------------------------------------------

def _read_sheet_rows(wb, sheet_name: str) -> list[dict]:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = list(rows[0])
    return [dict(zip(header, r)) for r in rows[1:]]


def test_phase4_scored_to_xlsx_emits_suppressed_sheet():
    """scored_to_xlsx now writes a `Suppressed` sub-sheet AND a
    `suppressed_by` column on the main Stage-1 drops sheet."""
    import audit_pack  # type: ignore
    from openpyxl import load_workbook

    with tempfile.TemporaryDirectory() as tmp:
        out_dir = Path(tmp) / "outputs"
        out_dir.mkdir(parents=True, exist_ok=True)
        scored = {
            "scan_date": "29991231",
            "stage1_passed": 1,
            "stage1_dropped": 3,
            "results": [],
            "triage_drops": [
                {"company": "Gamma", "title": "Sales Rep", "sector": "Banks",
                 "rule_reasons": ["no_strong_keywords"], "score": 1,
                 "hits_breakdown": {"strong": [], "medium": [], "weak": [],
                                    "level": []},
                 "link": "https://x.co/20", "source": "linkedin"},
                {"company": "RBC", "title": "Director ALM",
                 "sector": "Canadian Big 6 Banks",
                 "rule_reasons": ["suppressed_sector_60d"], "score": 7,
                 "hits_breakdown": {"strong": ["alm"], "medium": [], "weak": [],
                                    "level": []},
                 "link": "https://x.co/21", "source": "scrape"},
                {"company": "Acme", "title": "Risk Mgr",
                 "rule_reasons": ["suppressed_company_30d"], "score": 6,
                 "hits_breakdown": {"strong": [], "medium": [], "weak": [],
                                    "level": []},
                 "link": "https://x.co/22", "source": "linkedin"},
            ],
        }
        scored_path = out_dir / "worklist_scored.json"
        scored_path.write_text(json.dumps(scored), encoding="utf-8")

        data = audit_pack.scored_to_xlsx(scored_path)
        wb = load_workbook(io.BytesIO(data), read_only=True)
        assert "Suppressed" in wb.sheetnames

        s1_rows = _read_sheet_rows(wb, "Stage-1 drops")
        assert "suppressed_by" in s1_rows[0]
        # Non-suppressed row → empty string in the column
        gamma = next(r for r in s1_rows if r["company"] == "Gamma")
        assert gamma["suppressed_by"] == "" or gamma["suppressed_by"] is None
        # Sector mute → "sector_60d"
        rbc = next(r for r in s1_rows if r["company"] == "RBC")
        assert rbc["suppressed_by"] == "sector_60d"
        # Company mute → "company_30d"
        acme = next(r for r in s1_rows if r["company"] == "Acme")
        assert acme["suppressed_by"] == "company_30d"

        # Suppressed sub-sheet contains both muted rows, NOT the
        # no_strong_keywords drop.
        sup_rows = _read_sheet_rows(wb, "Suppressed")
        assert len(sup_rows) == 2
        sup_companies = {r["company"] for r in sup_rows}
        assert sup_companies == {"RBC", "Acme"}


def test_phase4_promote_to_xlsx_has_selection_mode_and_race_sheet():
    """promote_to_xlsx emits selection_mode columns + Suppressed (race) sheet
    + Run meta sheet."""
    import audit_pack  # type: ignore
    from openpyxl import load_workbook

    with tempfile.TemporaryDirectory() as tmp:
        out_dir = Path(tmp) / "outputs"
        out_dir.mkdir(parents=True, exist_ok=True)
        promote = {
            "scan_date": "29991231",
            "selection_mode": "manual",
            "min_score": 7,
            "include_watch": False,
            "summary": {"added": 2},
            "new_entries": [
                {"id": "auto-rbc-alm-9", "tier": 1, "score": 9,
                 "sector": "Canadian Big 6 Banks", "company": "RBC",
                 "title": "Director ALM",
                 "selection_mode": "manual_override_suppression",
                 "url": "https://x.co/21"},
                {"id": "auto-wave-treas-8", "tier": 2, "score": 8,
                 "sector": "Fintech", "company": "Wave",
                 "title": "Treasury",
                 "selection_mode": "manual",
                 "url": "https://x.co/30"},
            ],
            "skipped_rows": [
                {"reason": "geo", "company": "X", "title": "T",
                 "url": "https://x.co/99", "location": "NYC", "score": 8,
                 "verdict": "apply_now",
                 "selection_mode": "manual",
                 "note": "out-of-region"},
            ],
            "suppressed_after_score": [
                {"company": "TD", "title": "ALM Lead",
                 "sector": "Canadian Big 6 Banks", "score": 8,
                 "verdict": "apply_now",
                 "drop_reason": "suppressed_sector_60d",
                 "promoted_anyway": False, "selection_mode": "threshold",
                 "url": "https://x.co/40"},
            ],
        }
        promote_path = out_dir / "promote_report_29991231.json"
        promote_path.write_text(json.dumps(promote), encoding="utf-8")

        data = audit_pack.promote_to_xlsx(promote_path)
        wb = load_workbook(io.BytesIO(data), read_only=True)

        assert "Run" in wb.sheetnames
        assert "Promoted" in wb.sheetnames
        assert "Skipped" in wb.sheetnames
        assert "Suppressed (race)" in wb.sheetnames

        # Run meta carries the top-level selection_mode + counts
        run_rows = _read_sheet_rows(wb, "Run")
        meta = {r["key"]: r["value"] for r in run_rows}
        assert meta["selection_mode"] == "manual"
        assert meta["promoted_count"] == 2
        assert meta["suppressed_after_score_count"] == 1

        # Promoted sheet has selection_mode column with the override value
        p_rows = _read_sheet_rows(wb, "Promoted")
        assert "selection_mode" in p_rows[0]
        rbc = next(r for r in p_rows if r["company"] == "RBC")
        assert rbc["selection_mode"] == "manual_override_suppression"

        # Suppressed (race) sheet captures the dropped row
        sup_rows = _read_sheet_rows(wb, "Suppressed (race)")
        assert len(sup_rows) == 1
        assert sup_rows[0]["company"] == "TD"
        assert sup_rows[0]["drop_reason"] == "suppressed_sector_60d"


def test_phase4_build_audit_pack_includes_suppression_sheets():
    """End-to-end: full audit pack carries both the triage-suppressed and
    race-suppressed sheets, and the Summary sheet surfaces the counts."""
    import audit_pack  # type: ignore
    from openpyxl import load_workbook

    with tempfile.TemporaryDirectory() as tmp:
        synth_out = Path(tmp) / "outputs"
        _make_synth_outputs(synth_out, "29991231")
        # Replace scored + promote with envelopes that include suppression rows.
        scored = json.loads((synth_out / "worklist_scored.json")
                            .read_text(encoding="utf-8"))
        scored["triage_drops"].append({
            "company": "RBC", "title": "Director ALM",
            "sector": "Canadian Big 6 Banks",
            "rule_reasons": ["suppressed_sector_60d"], "score": 7,
            "hits_breakdown": {"strong": [], "medium": [], "weak": [],
                                "level": []},
            "link": "https://x.co/30", "source": "scrape",
        })
        (synth_out / "worklist_scored.json").write_text(
            json.dumps(scored), encoding="utf-8")
        promote = json.loads((synth_out / "promote_report_29991231.json")
                             .read_text(encoding="utf-8"))
        promote["selection_mode"] = "manual"
        promote["new_entries"][0]["selection_mode"] = "manual"
        promote["suppressed_after_score"] = [{
            "company": "TD", "title": "ALM Lead",
            "sector": "Canadian Big 6 Banks", "score": 8,
            "verdict": "apply_now",
            "drop_reason": "suppressed_sector_60d",
            "promoted_anyway": False, "selection_mode": "threshold",
            "url": "https://x.co/40",
        }]
        (synth_out / "promote_report_29991231.json").write_text(
            json.dumps(promote), encoding="utf-8")

        original_out = audit_pack.OUT_DIR
        audit_pack.OUT_DIR = synth_out
        try:
            data = audit_pack.build_audit_pack("29991231")
            wb = load_workbook(io.BytesIO(data), read_only=True)
            assert "Suppressed (triage)" in wb.sheetnames
            assert "Suppressed (race)" in wb.sheetnames

            triage_rows = _read_sheet_rows(wb, "Suppressed (triage)")
            assert any(r["company"] == "RBC" for r in triage_rows)

            race_rows = _read_sheet_rows(wb, "Suppressed (race)")
            assert any(r["company"] == "TD" for r in race_rows)

            # Summary sheet surfaces the counts
            summary_rows = _read_sheet_rows(wb, "Summary")
            metrics = {r.get("Metric"): r.get("Value")
                       for r in summary_rows
                       if r.get("Metric")}
            assert metrics.get("Suppressed (triage)") == 1
            assert metrics.get("Promote — suppressed_after_score") == 1
            assert metrics.get("Promote — selection_mode") == "manual"
        finally:
            audit_pack.OUT_DIR = original_out


def test_triage_to_xlsx_passed_dropped_suppressed_sheets():
    """triage_to_xlsx emits Passed / Dropped / Suppressed sheets with the
    suppressed_by column and the dedicated Suppressed sub-sheet."""
    import audit_pack  # type: ignore
    from openpyxl import load_workbook

    with tempfile.TemporaryDirectory() as tmp:
        out_dir = Path(tmp) / "outputs"
        out_dir.mkdir(parents=True, exist_ok=True)
        scored = {
            "scan_date": "29991231",
            "stage1_passed": 2, "stage1_dropped": 2,
            "results": [
                {"title": "Risk Mgr", "company": "Acme", "sector": "Banks",
                 "location": "Toronto", "source": "scrape",
                 "link": "https://x.co/1", "_triage": {"score": 5},
                 "fit": {"fit_verdict": "watch", "fit_score": 60}},
                {"title": "ALM", "company": "Beta", "sector": "Banks",
                 "location": "Toronto", "source": "scrape",
                 "link": "https://x.co/2", "_triage": {"score": 7},
                 "fit": {"fit_verdict": "apply_now", "fit_score": 80}},
            ],
            "triage_drops": [
                {"company": "Gamma", "title": "Sales", "sector": "Retail",
                 "rule_reasons": ["no_strong_keywords"], "score": 1,
                 "hits_breakdown": {}, "link": "https://x.co/3",
                 "source": "linkedin"},
                {"company": "RBC", "title": "Dir ALM",
                 "sector": "Canadian Big 6 Banks",
                 "rule_reasons": ["suppressed_sector_60d"], "score": 7,
                 "hits_breakdown": {}, "link": "https://x.co/4",
                 "source": "scrape"},
            ],
        }
        p = out_dir / "worklist_scored.json"
        p.write_text(json.dumps(scored), encoding="utf-8")

        data = audit_pack.triage_to_xlsx(p)
        wb = load_workbook(io.BytesIO(data), read_only=True)
        assert wb.sheetnames == ["Passed", "Dropped", "Suppressed"], wb.sheetnames

        passed = _read_sheet_rows(wb, "Passed")
        assert len(passed) == 2 and "stage1_score" in passed[0]
        dropped = _read_sheet_rows(wb, "Dropped")
        assert len(dropped) == 2 and "suppressed_by" in dropped[0]
        rbc = next(r for r in dropped if r["company"] == "RBC")
        assert rbc["suppressed_by"] == "sector_60d"
        sup = _read_sheet_rows(wb, "Suppressed")
        assert len(sup) == 1 and sup[0]["company"] == "RBC"
        print("[PASS] triage_to_xlsx: Passed/Dropped/Suppressed sheets correct")


def test_tracker_to_xlsx_one_sheet_per_status():
    """tracker_to_xlsx emits an `All` sheet plus one sheet per distinct
    status, with an `archived` column and Excel-safe sheet names."""
    import audit_pack  # type: ignore
    from openpyxl import load_workbook

    with tempfile.TemporaryDirectory() as tmp:
        tracker = {
            "jobs": [
                {"id": "a1", "company": "Acme", "title": "Risk Mgr",
                 "status": "Found", "tier": 1, "fit_score_numeric": 8,
                 "url": "https://x.co/1"},
                {"id": "a2", "company": "Beta", "title": "ALM",
                 "status": "Applied", "tier": 2, "fit_score_numeric": 7,
                 "date_applied": "2026-05-01", "url": "https://x.co/2"},
                {"id": "a3", "company": "Gamma", "title": "Quant",
                 "status": "Found", "tier": 1, "archived": True,
                 "url": "https://x.co/3"},
            ],
            "meta": {},
        }
        p = Path(tmp) / "job_tracker_data.json"
        p.write_text(json.dumps(tracker), encoding="utf-8")

        data = audit_pack.tracker_to_xlsx(p)
        wb = load_workbook(io.BytesIO(data), read_only=True)
        assert "All" in wb.sheetnames
        assert "Found" in wb.sheetnames and "Applied" in wb.sheetnames
        all_rows = _read_sheet_rows(wb, "All")
        assert len(all_rows) == 3 and "archived" in all_rows[0]
        found = _read_sheet_rows(wb, "Found")
        assert len(found) == 2  # both Found rows (incl. the archived one, tagged)
        # the archived Found row carries archived=True
        arch = [r for r in found if r["company"] == "Gamma"]
        assert arch and bool(arch[0]["archived"]) is True
        print("[PASS] tracker_to_xlsx: per-status sheets + archived column")


if __name__ == "__main__":
    test_audit_pack_round_trip()
    test_audit_pack_empty_legacy_scan()
    test_phase4_scored_to_xlsx_emits_suppressed_sheet()
    test_phase4_promote_to_xlsx_has_selection_mode_and_race_sheet()
    test_phase4_build_audit_pack_includes_suppression_sheets()
    test_triage_to_xlsx_passed_dropped_suppressed_sheets()
    test_tracker_to_xlsx_one_sheet_per_status()
    print("\nAll audit_pack tests passed.")
