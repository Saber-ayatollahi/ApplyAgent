"""Regression-locking tests for the findings confirmed by the round-1
adversarial review of the resume-pipeline / triage-split branch.

Unlike test_dedup_merge.py (which logs PASS/FAIL but only fails pytest on a
crash), every test here uses real `assert`s so a regression fails CI.

Covered behaviors (each maps to a confirmed review finding):
  - worklist disposition ledger reconciles (input == sum of buckets)
  - norm_url preserves identity query params (gh_jid) → distinct postings
  - degenerate-title rows on distinct LinkedIn ids are kept, not merged
  - misparse field-swap is REPAIRED (real company recovered), and the
    regression guard refuses to fabricate when the company is a role stem
  - parser-artifact rows are quarantined with reason 'parser_artifact'
  - no_url rows land in the no_url_dropped bucket
  - fit_scorer --triage-out writes a SEPARATE file, leaving _scored.json intact
  - audit_pack.worklist_to_xlsx builds the 4-sheet workbook off the new envelope
"""
from __future__ import annotations

import json
import sys
import tempfile
from pathlib import Path

import pytest

HERE = Path(__file__).resolve().parent
AUTO = HERE.parent
sys.path.insert(0, str(AUTO))

import worklist  # type: ignore
from worklist import norm_url, _is_degenerate_title, _normalize_title  # type: ignore


# ---------------------------------------------------------------------------
# norm_url — identity query-param preservation
# ---------------------------------------------------------------------------
def test_norm_url_preserves_gh_jid():
    """Two Greenhouse postings on the SAME path differing only by ?gh_jid=
    must canonicalize to DISTINCT keys (else one real job is lost to dedup)."""
    a = norm_url({"link": "https://boards.greenhouse.io/acme/jobs/search?gh_jid=111"})
    b = norm_url({"link": "https://boards.greenhouse.io/acme/jobs/search?gh_jid=222"})
    assert a != b, f"distinct gh_jid collapsed to one key: {a!r}"
    assert "gh_jid=111" in a and "gh_jid=222" in b


def test_norm_url_strips_tracking_but_keeps_identity():
    """Tracking noise (utm_*) is stripped; identity params survive. Two URLs
    differing ONLY by tracking must collapse to the same key."""
    a = norm_url({"link": "https://jobs.example.com/x/123?utm_source=li&gh_jid=9"})
    b = norm_url({"link": "https://jobs.example.com/x/123?gh_jid=9&utm_campaign=z"})
    assert a == b, f"tracking-only difference did not collapse: {a!r} vs {b!r}"
    assert "gh_jid=9" in a
    assert "utm" not in a


def test_norm_url_lever_path_disambiguates():
    """Lever encodes the id in the PATH, not a query param — two distinct
    Lever postings stay distinct without any ?lever= handling."""
    a = norm_url({"link": "https://jobs.lever.co/acme/abc-111"})
    b = norm_url({"link": "https://jobs.lever.co/acme/def-222"})
    assert a != b
    assert a == "https://jobs.lever.co/acme/abc-111"


# ---------------------------------------------------------------------------
# rebuild() — disposition ledger, repair, quarantine, reconciliation
# ---------------------------------------------------------------------------
@pytest.fixture
def redirected_outputs():
    """Redirect worklist module-level paths to a temp dir for the duration of
    the test, restoring them after."""
    tmp = Path(tempfile.mkdtemp(prefix="wl_review_"))
    saved = (worklist.OUT_DIR, worklist.WORKLIST,
             worklist.WORKLIST_SCORED, worklist.LEGACY_DIR)
    worklist.OUT_DIR = tmp
    worklist.WORKLIST = tmp / "worklist.json"
    worklist.WORKLIST_SCORED = tmp / "worklist_scored.json"
    worklist.LEGACY_DIR = tmp / "_legacy"
    try:
        yield tmp
    finally:
        (worklist.OUT_DIR, worklist.WORKLIST,
         worklist.WORKLIST_SCORED, worklist.LEGACY_DIR) = saved
        import shutil
        shutil.rmtree(tmp, ignore_errors=True)


def _write_inputs(tmp: Path, web_rows: list[dict], gmail_rows: list[dict]):
    from datetime import datetime
    today = datetime.now().strftime("%Y%m%d")
    (tmp / f"scan_{today}.json").write_text(
        json.dumps({"results": web_rows}), encoding="utf-8")
    (tmp / f"scan_gmail_{today}_120000.json").write_text(
        json.dumps({"results": gmail_rows}), encoding="utf-8")


def _rebuild_env(tmp: Path) -> dict:
    worklist.rebuild(quarantine=True)
    return json.loads((tmp / "worklist.json").read_text(encoding="utf-8"))


def test_reconciliation_balances(redirected_outputs):
    """Every input row lands in exactly one disposition bucket."""
    tmp = redirected_outputs
    web = [{"link": "https://jobs.td.com/job/1-Director-ALM",
            "company": "TD Bank", "title": "Director, ALM"}]
    gmail = [
        {"link": "https://ca.linkedin.com/jobs/view/analyst-rbc-100",
         "company": "RBC", "title": "Risk Analyst"},
        {"link": "https://ca.linkedin.com/jobs/view/mgr-cibc-200",
         "company": "CIBC", "title": "Treasury Manager"},
    ]
    _write_inputs(tmp, web, gmail)
    env = _rebuild_env(tmp)
    rec = env["reconciliation"]
    assert rec["balanced"] is True, f"reconciliation off by {rec['unaccounted']}: {rec}"
    assert rec["input"] == 3
    assert rec["accounted"] == rec["input"]


def test_degenerate_title_distinct_linkedin_kept(redirected_outputs):
    """Two DIFFERENT LinkedIn reqs at the same company sharing only a bare
    'Senior Manager' stem must BOTH survive (no false near-dup merge)."""
    tmp = redirected_outputs
    gmail = [
        {"link": "https://ca.linkedin.com/jobs/view/sm-rbc-111",
         "company": "RBC", "title": "Senior Manager"},
        {"link": "https://ca.linkedin.com/jobs/view/sm-rbc-222",
         "company": "RBC", "title": "Senior Manager"},
    ]
    _write_inputs(tmp, [], gmail)
    env = _rebuild_env(tmp)
    rbc = [r for r in env["results"] if r.get("company") == "RBC"]
    assert len(rbc) == 2, f"degenerate-title rows collapsed: {[r['link'] for r in rbc]}"
    assert env["reconciliation"]["balanced"] is True


def test_misparse_repair_recovers_real_company(redirected_outputs):
    """Field-swap corruption (company holds the role, title is
    '<role> <RealCo>') is REPAIRED, not discarded."""
    tmp = redirected_outputs
    # company='Treasury Manager' (NOT a degenerate stem), title echoes it + KOHO
    gmail = [{"link": "https://ca.linkedin.com/jobs/view/tm-koho-900",
              "company": "Treasury Manager",
              "title": "Treasury Manager KOHO",
              "location": "Toronto, ON"}]
    _write_inputs(tmp, [], gmail)
    env = _rebuild_env(tmp)
    assert env["reconciliation"]["repaired"] == 1, env["reconciliation"]
    row = env["results"][0]
    assert row["company"] == "KOHO", row
    assert row["title"] == "Treasury Manager", row


def test_misparse_repair_refuses_degenerate_company(redirected_outputs):
    """REGRESSION GUARD: when the company field is itself a bare role stem
    ('Senior Manager') prefixing a hyphenated title, that's a real hyphenated
    role — NOT a field-swap. The repair must NOT fire and fabricate a company."""
    tmp = redirected_outputs
    gmail = [{"link": "https://ca.linkedin.com/jobs/view/sm-amex-901",
              "company": "Senior Manager",
              "title": "Senior Manager Risk Management American Express",
              "location": "Toronto, ON"}]
    _write_inputs(tmp, [], gmail)
    env = _rebuild_env(tmp)
    # repair must NOT have fabricated 'Risk Management American Express'
    assert env["reconciliation"]["repaired"] == 0, env["reconciliation"]
    companies = [r.get("company") for r in env["results"]]
    assert "Risk Management American Express" not in companies, companies


def test_misparse_repair_refuses_midword_prefix(redirected_outputs):
    """ROUND-2 GUARD: company is a mid-word substring prefix of the title's
    first word ('Treasury Manager' ⊂ 'Treasury Managerial Group'). A true swap
    joins on whitespace; a mid-word cut would fabricate 'ial Group'. Refuse."""
    tmp = redirected_outputs
    gmail = [{"link": "https://ca.linkedin.com/jobs/view/midword-1",
              "company": "Treasury Manager",
              "title": "Treasury Managerial Group",
              "location": "Toronto, ON"}]
    _write_inputs(tmp, [], gmail)
    env = _rebuild_env(tmp)
    assert env["reconciliation"]["repaired"] == 0, env["reconciliation"]
    assert "ial Group" not in [r.get("company") for r in env["results"]]


def test_misparse_repair_refuses_dash_joined_role(redirected_outputs):
    """ROUND-2 GUARD: a hyphenated role ('Treasury Manager - Capital Markets')
    is NOT a field-swap (swaps join role→company on whitespace, not ' - ').
    Must not fabricate company='Capital Markets'."""
    tmp = redirected_outputs
    gmail = [{"link": "https://ca.linkedin.com/jobs/view/dashrole-1",
              "company": "Treasury Manager",
              "title": "Treasury Manager - Capital Markets",
              "location": "Toronto, ON"}]
    _write_inputs(tmp, [], gmail)
    env = _rebuild_env(tmp)
    assert env["reconciliation"]["repaired"] == 0, env["reconciliation"]
    assert "Capital Markets" not in [r.get("company") for r in env["results"]]


def test_artifact_row_quarantined(redirected_outputs):
    """LinkedIn section-header rows ('Jobs similar to …') are quarantined,
    never turned into a fake company."""
    tmp = redirected_outputs
    gmail = [{"link": "https://ca.linkedin.com/jobs/view/art-1",
              "company": "Jobs similar to Senior Risk Analyst",
              "title": "Jobs similar to Senior Risk Analyst"}]
    _write_inputs(tmp, [], gmail)
    env = _rebuild_env(tmp)
    reasons = [q.get("reason") for q in env.get("quarantine", [])]
    assert "parser_artifact" in reasons, env.get("quarantine")
    assert env["reconciliation"]["balanced"] is True


def test_no_url_row_bucketed(redirected_outputs):
    """A row with no canonical URL lands in no_url_dropped (was a silent drop)."""
    tmp = redirected_outputs
    web = [{"link": "", "company": "Ghost Co", "title": "No URL Role"}]
    _write_inputs(tmp, web, [])
    env = _rebuild_env(tmp)
    assert len(env.get("no_url_dropped", [])) == 1, env.get("no_url_dropped")
    assert env["reconciliation"]["no_url_dropped"] == 1
    assert env["reconciliation"]["balanced"] is True


# ---------------------------------------------------------------------------
# fit_scorer --triage-out — separate file, scores untouched
# ---------------------------------------------------------------------------
def test_triage_out_writes_separate_file_preserving_scores(monkeypatch):
    """fit_scorer --dry-run --triage-out writes the triage preview to the named
    file and leaves a pre-existing _scored.json byte-for-byte untouched."""
    import fit_scorer  # type: ignore
    tmp = Path(tempfile.mkdtemp(prefix="triageout_"))
    monkeypatch.setattr(fit_scorer, "OUT_DIR", tmp)
    try:
        scan = tmp / "pool.json"
        scan.write_text(json.dumps({"scan_date": "2026-06-02", "results": [
            {"link": "https://x/1", "company": "RBC", "title": "Risk Analyst"},
        ]}), encoding="utf-8")
        # Pre-existing real scores must survive a triage run.
        scored = tmp / "pool_scored.json"
        sentinel = {"stage2_scored": 42, "results": [{"fit": {"fit_score": 9}}]}
        scored.write_text(json.dumps(sentinel), encoding="utf-8")
        scored_before = scored.read_bytes()

        # main() reads sys.argv and parses with argparse.
        monkeypatch.setattr(sys, "argv", [
            "fit_scorer.py", "--scan", "pool.json", "--dry-run",
            "--triage-out", "pool_triage.json",
        ])
        rc = fit_scorer.main()

        assert rc == 0, f"triage dry-run exited {rc}"
        triage = tmp / "pool_triage.json"
        assert triage.exists(), "triage-out file not written"
        tj = json.loads(triage.read_text(encoding="utf-8"))
        assert tj.get("stage1_only") is True
        # The real scores file is untouched.
        assert scored.read_bytes() == scored_before, "triage clobbered _scored.json!"
    finally:
        import shutil
        shutil.rmtree(tmp, ignore_errors=True)


def test_dry_run_without_triage_out_uses_legacy_path(monkeypatch):
    """Legacy contract: --dry-run WITHOUT --triage-out still writes
    <scan>_scored.json (the 'Score dry-run' checkbox path)."""
    import fit_scorer  # type: ignore
    tmp = Path(tempfile.mkdtemp(prefix="triagelegacy_"))
    monkeypatch.setattr(fit_scorer, "OUT_DIR", tmp)
    try:
        (tmp / "pool.json").write_text(json.dumps({"scan_date": "2026-06-02",
            "results": [{"link": "https://x/1", "company": "RBC",
                         "title": "Risk Analyst"}]}), encoding="utf-8")
        monkeypatch.setattr(sys, "argv",
                            ["fit_scorer.py", "--scan", "pool.json", "--dry-run"])
        rc = fit_scorer.main()
        assert rc == 0
        legacy = tmp / "pool_scored.json"
        assert legacy.exists(), "legacy dry-run did not write _scored.json"
        assert json.loads(legacy.read_text(encoding="utf-8")).get("stage1_only") is True
    finally:
        import shutil
        shutil.rmtree(tmp, ignore_errors=True)


# ---------------------------------------------------------------------------
# audit_pack — 4-sheet workbook off the new envelope
# ---------------------------------------------------------------------------
def test_worklist_to_xlsx_four_sheets():
    """worklist_to_xlsx builds Summary/Raw/Merged/Dropped from the new
    disposition-ledger envelope without raising."""
    import audit_pack  # type: ignore
    from openpyxl import load_workbook
    import io

    tmp = Path(tempfile.mkdtemp(prefix="auditxlsx_"))
    try:
        env = {
            "rebuilt_at": "2026-06-02T10:00:00",
            "scan_date": "2026-06-02",
            "reconciliation": {
                "input": 5, "kept": 2, "merged": 1, "quarantined": 1,
                "geo_dropped": 0, "excluded": 1, "no_url_dropped": 0,
                "repaired": 1, "accounted": 5, "balanced": True, "unaccounted": 0,
            },
            "results": [
                {"company": "RBC", "title": "Risk Analyst", "link": "https://x/1"},
                {"company": "TD", "title": "Director, ALM", "link": "https://x/2"},
            ],
            "merged_pairs": [{
                "company": "RBC", "reason": "near_dup_company_title",
                "kept_title": "Risk Analyst", "dropped_title": "Sr Risk Analyst",
                "kept_url": "https://x/1", "dropped_url": "https://x/9",
                "kept_source": "scrape", "dropped_source": "gmail",
            }],
            "quarantine": [{"company": "Jobs similar to X", "title": "Jobs similar to X",
                            "link": "", "reason": "parser_artifact"}],
            "excluded": [{"company": "Scotiabank", "title": "Analyst",
                          "link": "https://x/3", "source": "scrape",
                          "reason": "excluded_company"}],
            "geo_dropped": [],
            "no_url_dropped": [],
        }
        p = tmp / "worklist.json"
        p.write_text(json.dumps(env), encoding="utf-8")
        data = audit_pack.worklist_to_xlsx(p)
        wb = load_workbook(io.BytesIO(data))
        assert set(wb.sheetnames) == {"Summary", "Raw", "Merged", "Dropped"}, wb.sheetnames
        # Merged sheet must carry the kept/dropped titles (the new fields).
        merged_vals = [c.value for row in wb["Merged"].iter_rows() for c in row]
        assert "Sr Risk Analyst" in merged_vals
        # Raw ledger covers every disposition that has rows.
        raw_vals = [c.value for row in wb["Raw"].iter_rows() for c in row]
        for disp in ("kept", "merged", "quarantined", "excluded"):
            assert disp in raw_vals, f"{disp} missing from Raw ledger"
    finally:
        import shutil
        shutil.rmtree(tmp, ignore_errors=True)


def test_worklist_to_xlsx_pre_ledger_backcompat():
    """A pre-ledger worklist.json (no reconciliation/excluded/geo keys) still
    produces a workbook — degrades gracefully."""
    import audit_pack  # type: ignore
    from openpyxl import load_workbook
    import io

    tmp = Path(tempfile.mkdtemp(prefix="auditold_"))
    try:
        env = {
            "results": [{"company": "RBC", "title": "Analyst", "link": "https://x/1"}],
            "merged_pairs": [],
        }
        p = tmp / "worklist.json"
        p.write_text(json.dumps(env), encoding="utf-8")
        data = audit_pack.worklist_to_xlsx(p)
        wb = load_workbook(io.BytesIO(data))
        assert "Summary" in wb.sheetnames
    finally:
        import shutil
        shutil.rmtree(tmp, ignore_errors=True)


# ---------------------------------------------------------------------------
# gmail_reader Mode E — trailing location-echo strip (round-2 fixes)
# ---------------------------------------------------------------------------
@pytest.mark.parametrize("title,company,location,expected", [
    # positive strip: full-location echo
    ("Analyst - Toronto, ON", "X", "Toronto, ON", "Analyst"),
    # city-part match (location has province, title has only city)
    ("Director - Toronto", "X", "Toronto, ON", "Director"),
    # genuine hyphenated role MUST be preserved (not a location)
    ("Manager - Customer Risk Analysis", "X", "Toronto, ON",
     "Manager - Customer Risk Analysis"),
    ("EHS Manager - Canada & HQ/DCs", "Acme", "Toronto, ON",
     "EHS Manager - Canada & HQ/DCs"),
    # multi-dash: strip only the trailing echo segment
    ("Senior Manager - Risk - Toronto, ON", "X", "Toronto, ON",
     "Senior Manager - Risk"),
    # hyphenated city name (echo segment itself contains a dash)
    ("Analyst - Winston-Salem, NC", "X", "Winston-Salem, NC", "Analyst"),
    # Mode D stripped a city-equals-company token first; clean the dangling dash
    ("Senior Analyst - Toronto", "Toronto", "Toronto, ON", "Senior Analyst"),
    # empty location: nothing to match, leave title alone
    ("Analyst - Toronto, ON", "X", "", "Analyst - Toronto, ON"),
])
def test_mode_e_location_echo(title, company, location, expected):
    from gmail_reader import _clean_alert_fields  # type: ignore
    got_title, _, _ = _clean_alert_fields(title, company, location)
    assert got_title == expected, f"{title!r} -> {got_title!r} (expected {expected!r})"


# ---------------------------------------------------------------------------
# run_pipeline --triage-out guards (round-1 guards, round-2 test lock)
# ---------------------------------------------------------------------------
def test_run_pipeline_triage_out_requires_dry_run(monkeypatch):
    """--triage-out without --score-dry-run must be refused (exit 2) before any
    side effect."""
    import run_pipeline  # type: ignore
    monkeypatch.setattr(sys, "argv", [
        "run_pipeline.py", "--skip-scrape", "--skip-promote",
        "--triage-out", "x.json",
    ])
    assert run_pipeline.main() == 2


def test_run_pipeline_triage_out_refuses_promote(monkeypatch):
    """--triage-out with promote enabled (no --skip-promote) must be refused —
    promoting would silently use a stale worklist_scored.json."""
    import run_pipeline  # type: ignore
    monkeypatch.setattr(sys, "argv", [
        "run_pipeline.py", "--skip-scrape", "--score-dry-run",
        "--triage-out", "x.json",
    ])
    assert run_pipeline.main() == 2
